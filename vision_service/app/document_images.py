from __future__ import annotations

import io
import zipfile
from pathlib import Path

from PIL import Image


SUPPORTED_DOCUMENT_EXTENSIONS = {".pdf", ".docx", ".xlsx"}


def _image_dimensions(data: bytes) -> tuple[int | None, int | None]:
    try:
        with Image.open(io.BytesIO(data)) as image:
            return image.width, image.height
    except Exception:
        return None, None


def _extension_from_format(fmt: str | None) -> str:
    value = (fmt or "png").lower().strip(".")
    if value == "jpeg":
        value = "jpg"
    return f".{value}"


def extract_pdf_images(
    file_bytes: bytes,
    min_width: int,
    min_height: int,
) -> list[dict]:
    import fitz

    document = fitz.open(stream=file_bytes, filetype="pdf")
    results: list[dict] = []
    number = 0

    try:
        for page_index in range(document.page_count):
            page = document.load_page(page_index)
            seen_on_page: set[int] = set()

            for image_info in page.get_images(full=True):
                xref = int(image_info[0])
                if xref in seen_on_page:
                    continue
                seen_on_page.add(xref)

                extracted = document.extract_image(xref)
                data = extracted.get("image")
                if not data:
                    continue

                width = extracted.get("width")
                height = extracted.get("height")
                if width is None or height is None:
                    width, height = _image_dimensions(data)

                if (
                    width is not None
                    and height is not None
                    and (width < min_width or height < min_height)
                ):
                    continue

                number += 1
                results.append(
                    {
                        "image_number": number,
                        "source": "pdf_embedded_image",
                        "page": page_index + 1,
                        "sheet": None,
                        "cell": None,
                        "media_name": f"xref_{xref}",
                        "extension": _extension_from_format(
                            extracted.get("ext")
                        ),
                        "width": width,
                        "height": height,
                        "bytes": data,
                    }
                )
    finally:
        document.close()

    return results


def extract_pdf_native_text(file_bytes: bytes) -> list[dict]:
    """
    Texto ya presente en el PDF, sin pasar por OCR.

    Un PDF exportado desde Word o Google Docs lleva el texto real dentro.
    Transcribirlo por OCR sería introducir errores donde no los hay: la capa
    nativa es exacta por definición. Solo se recurre a la visión para lo que
    de verdad es una imagen.
    """
    import fitz

    document = fitz.open(stream=file_bytes, filetype="pdf")
    pages: list[dict] = []

    try:
        for index in range(document.page_count):
            text = document.load_page(index).get_text().strip()
            if text:
                pages.append(
                    {
                        "page": index + 1,
                        "characters": len(text),
                        "text": text,
                    }
                )
    finally:
        document.close()

    return pages


def render_pdf_pages(
    file_bytes: bytes,
    dpi: int = 200,
    max_pages: int = 20,
) -> list[dict]:
    """
    Renderiza páginas completas como imagen.

    Reserva para el PDF que no trae ni texto nativo ni imágenes incrustadas
    utilizables: por ejemplo un escaneo vectorizado, o una página donde la
    tarea está dibujada en lugar de insertada como foto.
    """
    import fitz

    document = fitz.open(stream=file_bytes, filetype="pdf")
    results: list[dict] = []
    zoom = dpi / 72.0

    try:
        for index in range(min(document.page_count, max_pages)):
            pixmap = document.load_page(index).get_pixmap(
                matrix=fitz.Matrix(zoom, zoom)
            )
            results.append(
                {
                    "image_number": index + 1,
                    "source": "pdf_rendered_page",
                    "page": index + 1,
                    "sheet": None,
                    "cell": None,
                    "media_name": f"page_{index + 1}",
                    "extension": ".png",
                    "width": pixmap.width,
                    "height": pixmap.height,
                    "bytes": pixmap.tobytes("png"),
                }
            )
    finally:
        document.close()

    return results


def extract_docx_images(
    file_bytes: bytes,
    min_width: int,
    min_height: int,
) -> list[dict]:
    results: list[dict] = []

    with zipfile.ZipFile(io.BytesIO(file_bytes)) as archive:
        media = sorted(
            name
            for name in archive.namelist()
            if name.startswith("word/media/")
            and not name.endswith("/")
        )

        for name in media:
            data = archive.read(name)
            width, height = _image_dimensions(data)

            if (
                width is not None
                and height is not None
                and (width < min_width or height < min_height)
            ):
                continue

            extension = Path(name).suffix.lower() or ".png"
            results.append(
                {
                    "image_number": len(results) + 1,
                    "source": "docx_embedded_image",
                    "page": None,
                    "sheet": None,
                    "cell": None,
                    "media_name": Path(name).name,
                    "extension": extension,
                    "width": width,
                    "height": height,
                    "bytes": data,
                }
            )

    return results


def extract_xlsx_images(
    file_bytes: bytes,
    min_width: int,
    min_height: int,
) -> list[dict]:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    workbook = load_workbook(
        io.BytesIO(file_bytes),
        data_only=True,
    )
    results: list[dict] = []

    try:
        for worksheet in workbook.worksheets:
            for image in getattr(worksheet, "_images", []):
                data = image._data()
                width, height = _image_dimensions(data)

                if (
                    width is not None
                    and height is not None
                    and (width < min_width or height < min_height)
                ):
                    continue

                cell = None
                try:
                    row = image.anchor._from.row + 1
                    col = image.anchor._from.col + 1
                    cell = f"{get_column_letter(col)}{row}"
                except Exception:
                    pass

                extension = ".png"
                fmt = getattr(image, "format", None)
                if fmt:
                    extension = _extension_from_format(fmt)

                results.append(
                    {
                        "image_number": len(results) + 1,
                        "source": "xlsx_embedded_image",
                        "page": None,
                        "sheet": worksheet.title,
                        "cell": cell,
                        "media_name": None,
                        "extension": extension,
                        "width": width,
                        "height": height,
                        "bytes": data,
                    }
                )
    finally:
        workbook.close()

    return results


def extract_document_images(
    filename: str,
    file_bytes: bytes,
    min_width: int,
    min_height: int,
) -> list[dict]:
    extension = Path(filename).suffix.lower()

    if extension == ".pdf":
        return extract_pdf_images(
            file_bytes,
            min_width,
            min_height,
        )
    if extension == ".docx":
        return extract_docx_images(
            file_bytes,
            min_width,
            min_height,
        )
    if extension == ".xlsx":
        return extract_xlsx_images(
            file_bytes,
            min_width,
            min_height,
        )

    raise ValueError(
        "Documento no soportado. Usa PDF, DOCX o XLSX."
    )
